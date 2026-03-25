"""
Excel 导出服务。

这个模块专门负责把统一财务报告导出为 `.xlsx` 文件。
为了保持口径一致，Excel 所有数据都来自 report_service 的统一汇总结果。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from core.config import settings
from schemas.export import ExportFileResponse
from schemas.report import FinancialReportResponse
from services.reporting.report_service import (
    build_financial_objective_summary,
    generate_financial_report,
)


def _ensure_exports_dir() -> Path:
    """
    确保导出目录存在。
    """
    export_dir = Path(settings.EXPORTS_DIR)
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _style_header(row) -> None:
    """
    统一设置表头样式。
    """
    for cell in row:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def export_financial_excel(
    db: Session,
    memory_db: Session,
    period: str,
    report: FinancialReportResponse | None = None,
    objective_data: dict[str, Any] | None = None,
    model: str | None = None,
) -> ExportFileResponse:
    """
    生成指定期间的 Excel 财务包。

    工作簿结构：
    - Summary: 关键指标和账龄总览
    - Vouchers: 凭证主表
    - VoucherEntries: 凭证明细分录
    - Receivables: 应收应付清单
    - AI_Report: AI 财务报告正文

    参数设计说明：
    - 普通路由调用时可以只传 period，函数内部会自行生成报告。
    - 如果上游已经生成过 report / objective_data，可以直接复用，
      避免重复查库和重复调用模型。
    """
    if objective_data is None:
        objective_data = build_financial_objective_summary(db, period)

    if report is None:
        report = generate_financial_report(db, memory_db, period, model=model)

    vouchers = objective_data["vouchers"]
    receivables = objective_data["receivables"]

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.freeze_panes = "A2"

    summary_sheet.append(["指标", "数值"])
    _style_header(summary_sheet[1])
    for metric in report.metrics:
        summary_sheet.append([metric.label, float(metric.value)])
        summary_sheet[f"B{summary_sheet.max_row}"].number_format = '#,##0.00;(#,##0.00);-'

    summary_sheet.append([])
    summary_sheet.append(["应收账龄", "金额"])
    _style_header(summary_sheet[summary_sheet.max_row])
    for item in report.objective_summary.receivable_aging:
        summary_sheet.append([item.label, float(item.amount)])
        summary_sheet[f"B{summary_sheet.max_row}"].number_format = '#,##0.00;(#,##0.00);-'

    summary_sheet.append([])
    summary_sheet.append(["应付账龄", "金额"])
    _style_header(summary_sheet[summary_sheet.max_row])
    for item in report.objective_summary.payable_aging:
        summary_sheet.append([item.label, float(item.amount)])
        summary_sheet[f"B{summary_sheet.max_row}"].number_format = '#,##0.00;(#,##0.00);-'

    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 18

    voucher_sheet = workbook.create_sheet("Vouchers")
    voucher_sheet.freeze_panes = "A2"
    voucher_sheet.append(["ID", "日期", "摘要", "状态", "创建人", "创建时间"])
    _style_header(voucher_sheet[1])
    for voucher in vouchers:
        voucher_sheet.append([
            voucher.id,
            voucher.date.isoformat(),
            voucher.memo,
            voucher.status.value,
            voucher.created_by,
            voucher.created_at.isoformat(),
        ])

    entry_sheet = workbook.create_sheet("VoucherEntries")
    entry_sheet.freeze_panes = "A2"
    entry_sheet.append(["凭证ID", "分录ID", "科目", "借方", "贷方"])
    _style_header(entry_sheet[1])
    for voucher in vouchers:
        for entry in voucher.entries:
            entry_sheet.append([
                voucher.id,
                entry.id,
                entry.account.name if entry.account else "",
                float(entry.debit),
                float(entry.credit),
            ])
            entry_sheet[f"D{entry_sheet.max_row}"].number_format = '#,##0.00;(#,##0.00);-'
            entry_sheet[f"E{entry_sheet.max_row}"].number_format = '#,##0.00;(#,##0.00);-'

    receivable_sheet = workbook.create_sheet("Receivables")
    receivable_sheet.freeze_panes = "A2"
    receivable_sheet.append(["ID", "类型", "往来单位", "金额", "到期日", "状态", "备注"])
    _style_header(receivable_sheet[1])
    for item in receivables:
        receivable_sheet.append([
            item.id,
            item.type.value,
            item.party,
            float(item.amount),
            item.due_date.isoformat() if item.due_date else "",
            item.status.value,
            item.memo or "",
        ])
        receivable_sheet[f"D{receivable_sheet.max_row}"].number_format = (
            '#,##0.00;(#,##0.00);-'
        )

    ai_sheet = workbook.create_sheet("AI_Report")
    ai_sheet["A1"] = "财务概览"
    ai_sheet["A2"] = report.narrative.executive_summary
    ai_sheet["A4"] = "财务分析"
    ai_sheet["A5"] = report.narrative.analysis
    ai_sheet["A7"] = "风险与异常"
    ai_sheet["A8"] = "\n".join(report.narrative.risks) if report.narrative.risks else "暂无"
    ai_sheet["A10"] = "建议动作"
    ai_sheet["A11"] = (
        "\n".join(report.narrative.recommendations)
        if report.narrative.recommendations
        else "暂无"
    )
    ai_sheet.column_dimensions["A"].width = 100
    for row_index in [2, 5, 8, 11]:
        ai_sheet[f"A{row_index}"].alignment = Alignment(wrap_text=True, vertical="top")

    # 给各个工作表补齐较合理的列宽，避免打开后第一眼就出现内容挤压。
    for worksheet in [voucher_sheet, entry_sheet, receivable_sheet]:
        for column_letter, width in {
            "A": 14,
            "B": 18,
            "C": 24,
            "D": 18,
            "E": 18,
            "F": 24,
            "G": 30,
        }.items():
            worksheet.column_dimensions[column_letter].width = width

    export_dir = _ensure_exports_dir()
    filename = f"fina-report-{period}.xlsx"
    output_path = export_dir / filename
    workbook.save(output_path)

    return ExportFileResponse(
        file_type="xlsx",
        period=period,
        filename=filename,
        path=str(output_path.resolve()),
    )
